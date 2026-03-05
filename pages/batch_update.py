import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import xlrd
from xlutils.copy import copy as xl_copy
from utils.excel_helpers import load_excel, get_headers_only, find_col_index, to_xlsx_stream, queue_download_js

def render_batch_update():
    st.markdown('<div class="upload-card">', unsafe_allow_html=True)
    up_c1, up_c2 = st.columns([1.2, 2])
    with up_c1:
        st.markdown("##### 📄 A表 (来源数据)")
        f_a = st.file_uploader("A", type=['xlsx', 'xls'], label_visibility="collapsed", key="ua")
        c_as, c_ac = st.columns(2)
        has = c_as.number_input("起行", 0, 20, 0, key="has_global")
        hac = c_ac.number_input("行数", 1, 5, 1, key="hac_global")
    with up_c2:
        st.markdown("##### 📑 B表 (目标多选)")
        fs_b = st.file_uploader("B", type=['xlsx', 'xls'], accept_multiple_files=True, label_visibility="collapsed", key="ubs")
        c_bs, c_bc = st.columns(2)
        hbs = c_bs.number_input("起行", 0, 20, 0, key="hbs_global")
        hbc = c_bc.number_input("行数", 1, 5, 1, key="hbc_global")
    st.markdown('</div>', unsafe_allow_html=True)

    if f_a and fs_b:
        df_a_g = load_excel(f_a, has, hac)
        df_b_g = load_excel(fs_b[0], hbs, hbc)

        if df_a_g is not None and df_b_g is not None:
            st.markdown('<div class="config-card">', unsafe_allow_html=True)
            st.caption("⚙️ 全局默认配置 (修改后将同步至下方个性化默认值)")
            cols = st.columns(4)
            ak = cols[0].selectbox("A匹配列", df_a_g.columns, key="gak")
            av = cols[1].selectbox("A取值列", df_a_g.columns, key="gav")
            bk = cols[2].selectbox("B匹配列", df_b_g.columns, key="gbk")
            bt = cols[3].selectbox("B更新列", df_b_g.columns, key="gbt")
            st.markdown('</div>', unsafe_allow_html=True)

            overrides = {}
            with st.expander(f"🔍 个性化清单 ({len(fs_b)}份)"):
                for i, fb in enumerate(fs_b):
                    oc1, oc2 = st.columns([5, 1])
                    oc1.caption(f"📄 {fb.name}")
                    # 只要勾选了个性化，才记录 override
                    if oc2.checkbox("个性化", key=f"is_p_{i}"):
                        st.markdown('<div class="personal-box">', unsafe_allow_html=True)
                        r1, r2, r3, r4 = st.columns(4)
                        p_as = r1.number_input("A起", 0, 20, value=has, key=f"pas_{i}")
                        p_ac = r2.number_input("A行", 1, 5, value=hac, key=f"pac_{i}")
                        p_bs = r3.number_input("B起", 0, 20, value=hbs, key=f"pbs_{i}")
                        p_bc = r4.number_input("B行", 1, 5, value=hbc, key=f"pbc_{i}")
                        
                        df_al, df_bl = load_excel(f_a, p_as, p_ac), load_excel(fb, p_bs, p_bc)
                        if df_al is not None and df_bl is not None:
                            def get_safe_idx(options, target):
                                try: return list(options).index(target)
                                except: return 0
                            l1, l2, l3, l4 = st.columns(4)
                            p_ak = l1.selectbox("A匹", df_al.columns, index=get_safe_idx(df_al.columns, ak), key=f"pak_{i}")
                            p_av = l2.selectbox("A取", df_al.columns, index=get_safe_idx(df_al.columns, av), key=f"pav_{i}")
                            p_bk = l3.selectbox("B匹", df_bl.columns, index=get_safe_idx(df_bl.columns, bk), key=f"pbk_{i}")
                            p_bt = l4.selectbox("B更", df_bl.columns, index=get_safe_idx(df_bl.columns, bt), key=f"pbt_{i}")
                            overrides[i] = {'has':p_as,'hac':p_ac,'hbs':p_bs,'hbc':p_bc,'ak':p_ak,'av':p_av,'bk':p_bk,'bt':p_bt}
                        st.markdown('</div>', unsafe_allow_html=True)

            if st.button("🚀 开始批量处理", use_container_width=True):
                temp_results = []
                error_logs = []
                prog = st.progress(0)
                for i, fb in enumerate(fs_b):
                    try:
                        conf = overrides.get(i, {'has':has,'hac':hac,'hbs':hbs,'hbc':hbc,'ak':ak,'av':av,'bk':bk,'bt':bt})
                        df_ca = load_excel(f_a, conf['has'], conf['hac'])
                        mapping = dict(zip(df_ca[conf['ak']].astype(str).str.strip(), df_ca[conf['av']]))
                        
                        headers = get_headers_only(fb, conf['hbs'], conf['hbc'])
                        ik, it = find_col_index(conf['bk'], headers), find_col_index(conf['bt'], headers)
                        
                        if ik != -1 and it != -1:
                            fb.seek(0)
                            file_ext = fb.name.rsplit('.', 1)[-1].lower()
                            
                            if file_ext == 'xls':
                                rb = xlrd.open_workbook(file_contents=fb.read(), formatting_info=True)
                                wb = xl_copy(rb)
                                ws = wb.get_sheet(0)
                                sheet_read = rb.sheet_by_index(0)
                                
                                col_k_idx = ik - 1
                                col_t_idx = it - 1
                                start_row_idx = conf['hbs'] + conf['hbc']
                                
                                for r in range(start_row_idx, sheet_read.nrows):
                                    cell_val = sheet_read.cell_value(r, col_k_idx)
                                    kv = str(cell_val).strip()
                                    if kv in mapping:
                                        ws.write(r, col_t_idx, mapping[kv])
                                
                                out = BytesIO()
                                wb.save(out)
                                temp_results.append((fb.name, out.getvalue()))
                            else:
                                wb = openpyxl.load_workbook(fb)
                                ws = wb.active
                                h_row = conf['hbs'] + conf['hbc']
                                for r in range(h_row + 1, ws.max_row + 1):
                                    kv = str(ws.cell(r, ik).value or "").strip()
                                    if kv in mapping: ws.cell(r, it).value = mapping[kv]
                                out = BytesIO(); wb.save(out)
                                temp_results.append((fb.name.rsplit('.', 1)[0] + ".xlsx", out.getvalue()))
                        else:
                            missing = []
                            if ik == -1: missing.append(f"匹配列 '{conf['bk']}'")
                            if it == -1: missing.append(f"更新列 '{conf['bt']}'")
                            error_logs.append(f"❌ {fb.name}: 未找到 {' 和 '.join(missing)}")
                    except Exception as e:
                        error_logs.append(f"⚠️ {fb.name}: 处理出错 - {str(e)}")
                    prog.progress((i + 1) / len(fs_b))
                
                st.session_state.batch_results = temp_results
                
                if error_logs:
                    with st.expander("🚨 处理异常报告", expanded=True):
                        for log in error_logs:
                            st.write(log)

            if st.session_state.batch_results:
                res = st.session_state.batch_results
                st.markdown(f'<div class="download-bar"><span>✅ 处理完成 ({len(res)}个)</span></div>', unsafe_allow_html=True)
                if st.button(f"📥 按顺序自动下载全部 {len(res)} 个文件", use_container_width=True, type="primary"):
                    st.components.v1.html(queue_download_js(res), height=0)
                    st.toast("正在启动队列下载，请允许浏览器下载多个文件...", icon="⌛")
