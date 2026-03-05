import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from copy import copy as obj_copy
from utils.excel_helpers import load_excel, get_headers_only, to_xlsx_stream

def render_field_extraction():
    st.markdown('<div class="upload-card">', unsafe_allow_html=True)
    st.markdown("##### 🔍 单文件字段提取")
    f_ex = st.file_uploader("提取文件", type=['xlsx', 'xls'], accept_multiple_files=False, label_visibility="collapsed", key="uex")
    
    if f_ex:
        st.markdown('<div class="config-card">', unsafe_allow_html=True)
        st.markdown("##### ⚙️ 提取配置")
        c1, c2 = st.columns(2)
        exs = c1.number_input("表头起行", 0, 10, 0, key="exs")
        exc = c2.number_input("表头行数", 1, 5, 1, key="exc")
        
        df = load_excel(f_ex, exs, exc)
        if df is not None:
            sel_cols = st.multiselect("选择需要保留的字段：", options=df.columns, default=[], key="sel_cols")
            
            st.markdown("##### 📝 导出配置")
            default_name_tpl = f_ex.name.rsplit('.', 1)[0] + "_提取"
            ex_name = st.text_input("导出文件名", value=default_name_tpl, key="ex_name")
            st.markdown('</div>', unsafe_allow_html=True)

            if st.button("🚀 执行提取", use_container_width=True, key="run_ex"):
                if not sel_cols:
                    st.error("⚠️ 请至少选择一个字段")
                else:
                    try:
                        f_ex.seek(0)
                        if f_ex.name.lower().endswith('.xls'):
                            stream = to_xlsx_stream(f_ex)
                        else:
                            stream = f_ex
                        
                        if stream is None:
                            st.error(f"❌ {f_ex.name}: 文件读取失败")
                        else:
                            wb = openpyxl.load_workbook(stream)
                            ws = wb.active
                            file_headers = get_headers_only(f_ex, exs, exc)
                            
                            col_map = {}
                            for idx, h in enumerate(file_headers):
                                if h in col_map: continue
                                col_map[h] = idx + 1

                            valid_sel_cols = [c for c in sel_cols if c in col_map]
                            
                            if not valid_sel_cols:
                                st.error("⚠️ 未找到匹配的列")
                            else:
                                new_wb = openpyxl.Workbook()
                                new_ws = new_wb.active
                                
                                max_row = ws.max_row
                                src_col_indices = [col_map[c] for c in valid_sel_cols]
                                
                                for row_idx in range(1, max_row + 1):
                                    for new_col_idx, src_col_idx in enumerate(src_col_indices, start=1):
                                        source_cell = ws.cell(row=row_idx, column=src_col_idx)
                                        target_cell = new_ws.cell(row=row_idx, column=new_col_idx)
                                        
                                        target_cell.value = source_cell.value
                                        
                                        if source_cell.has_style:
                                            target_cell.font = obj_copy(source_cell.font)
                                            target_cell.border = obj_copy(source_cell.border)
                                            target_cell.fill = obj_copy(source_cell.fill)
                                            target_cell.number_format = obj_copy(source_cell.number_format)
                                            target_cell.protection = obj_copy(source_cell.protection)
                                            target_cell.alignment = obj_copy(source_cell.alignment)
                                
                                for i, col_cells in enumerate(new_ws.columns, start=1):
                                    length = 0
                                    for cell in col_cells:
                                        if cell.value:
                                            val_str = str(cell.value)
                                            line_len = 0
                                            for char in val_str:
                                                if '\u4e00' <= char <= '\u9fff':
                                                    line_len += 2
                                                else:
                                                    line_len += 1
                                            length = max(length, line_len)
                                    
                                    adjusted_width = min(length + 2, 50) 
                                    col_letter = openpyxl.utils.get_column_letter(i)
                                    new_ws.column_dimensions[col_letter].width = adjusted_width

                                out = BytesIO()
                                new_wb.save(out)
                                
                                st.success("✅ 提取完成")
                                st.download_button(
                                    label="📥 下载提取文件",
                                    data=out.getvalue(),
                                    file_name=f"{ex_name}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary",
                                    use_container_width=True
                                )
                    except Exception as e:
                        st.error(f"⚠️ 处理出错: {str(e)}")
