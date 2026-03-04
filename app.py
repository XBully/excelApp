import streamlit as st
import pandas as pd
import openpyxl
import os
from io import BytesIO

# --- 页面配置 ---
st.set_page_config(
    page_title="小雷同学的Excel处理小助手",
    page_icon="💗",
    layout="wide"
)

# ======================================================
# 通用函数：读取 Excel（兼容 xls / xlsx）
# ======================================================
def load_excel(uploaded_file, header_rows):
    try:
        engine = 'xlrd' if uploaded_file.name.endswith('.xls') else 'openpyxl'
        return pd.read_excel(uploaded_file, header=header_rows, engine=engine)
    except Exception as e:
        st.error(f"读取失败: {e}")
        return None


# ======================================================
# 侧边栏
# ======================================================
with st.sidebar:
    st.title("🚀 功能选择")
    mode = st.radio(
        "请选择操作：",
        ["🔄 关联更新 (两表匹配)", "✨ 自定义提取 (单表筛选)"]
    )
    st.divider()
    st.caption("Excel 全能助手 v1.0")
    st.caption("支持多行表头 + xls / xlsx")


# ======================================================
# 模式一：关联更新
# ======================================================
if mode == "🔄 关联更新 (两表匹配)":

    st.header("🔄 关联更新 (VLOOKUP 模式)")
    st.info("根据 A 表数据更新 B 表指定列，保持 B 表原始格式")

    col_a, col_b = st.columns(2)

    with col_a:
        st.subheader("1️⃣ 上传 A 表")
        file_a = st.file_uploader("上传 A 表", type=['xlsx', 'xls'], key="update_a")
        h_a = st.number_input("A 表表头行", min_value=0, value=0)

    with col_b:
        st.subheader("2️⃣ 上传 B 表")
        file_b = st.file_uploader("上传 B 表", type=['xlsx', 'xls'], key="update_b")
        h_b = st.number_input("B 表表头行", min_value=0, value=0)

    if file_a and file_b:
        df_a = load_excel(file_a, h_a)
        df_b_preview = load_excel(file_b, h_b)

        if df_a is not None and df_b_preview is not None:

            st.divider()
            c1, c2, c3 = st.columns(3)

            with c1:
                a_key = st.selectbox("A表：匹配列", df_a.columns)
                a_val = st.selectbox("A表：取值列", df_a.columns)

            with c2:
                b_key = st.selectbox("B表：匹配列", df_b_preview.columns)
                b_target = st.selectbox("B表：目标更新列", df_b_preview.columns)

            with c3:
                default_out_name = os.path.splitext(file_b.name)[0]
                out_name = st.text_input("导出文件名:", value=default_out_name)

            if st.button("🚀 执行关联更新"):
                try:
                    mapping = dict(
                        zip(
                            df_a[a_key].astype(str).str.strip(),
                            df_a[a_val]
                        )
                    )

                    file_b.seek(0)

                    if file_b.name.endswith('.xls'):
                        df_full = pd.read_excel(file_b, header=None, engine='xlrd')
                        temp_io = BytesIO()
                        df_full.to_excel(temp_io, index=False, header=False, engine='openpyxl')
                        temp_io.seek(0)
                        wb = openpyxl.load_workbook(temp_io)
                    else:
                        wb = openpyxl.load_workbook(file_b)

                    ws = wb.active
                    headers = [str(cell.value) for cell in ws[h_b + 1]]

                    idx_id = headers.index(str(b_key)) + 1
                    idx_target = headers.index(str(b_target)) + 1

                    for r in range(h_b + 2, ws.max_row + 1):
                        raw_id = ws.cell(row=r, column=idx_id).value
                        clean_id = str(raw_id).strip() if raw_id else ""
                        if clean_id in mapping:
                            ws.cell(row=r, column=idx_target).value = mapping[clean_id]

                    out_io = BytesIO()
                    wb.save(out_io)

                    st.success("✅ 更新成功！")
                    st.download_button(
                        f"📥 下载 {out_name}.xlsx",
                        out_io.getvalue(),
                        f"{out_name}.xlsx"
                    )

                except Exception as e:
                    st.error(f"错误: {e}")


# ======================================================
# 模式二：自定义提取（支持多行表头）
# ======================================================
elif mode == "✨ 自定义提取 (单表筛选)":

    st.header("✨ 自定义提取 (支持跨行表头)")
    st.info("支持 1~N 行表头，自动合并为可勾选字段")

    file_extract = st.file_uploader(
        "上传原始 Excel 文件",
        type=['xlsx', 'xls'],
        key="extract_file"
    )

    if file_extract:

        h_ex = st.number_input("表头起始行(从0开始)", min_value=0, value=0)
        header_count = st.number_input("表头行数", min_value=1, value=1)

        try:
            engine = 'xlrd' if file_extract.name.endswith('.xls') else 'openpyxl'

            df_ex = pd.read_excel(
                file_extract,
                header=list(range(h_ex, h_ex + header_count)),
                engine=engine
            )

            # ---------- 多级表头处理 ----------
            # original_columns = df_ex.columns  # 备份原始表头结构 (已移除恢复逻辑，暂不需要)
            flat_columns = []

            if isinstance(df_ex.columns, pd.MultiIndex):
                for col in df_ex.columns:
                    col_parts = [str(i) for i in col if str(i) != "nan" and "Unnamed" not in str(i)]
                    flat_columns.append(" - ".join(col_parts))
            else:
                flat_columns = df_ex.columns.astype(str).tolist()
            
            # --- 表头去重处理 ---
            # Streamlit/PyArrow 不支持重复列名，需进行重命名处理
            deduped_columns = []
            seen = {}
            for col in flat_columns:
                if col in seen:
                    seen[col] += 1
                    new_col = f"{col}_{seen[col]}"
                    deduped_columns.append(new_col)
                else:
                    seen[col] = 0
                    deduped_columns.append(col)
            flat_columns = deduped_columns
            
            # 创建映射关系 (如果后续需要恢复原始表头可启用，目前仅用于扁平化导出)
            # col_mapping = dict(zip(flat_columns, original_columns))
            
            # 更新 DataFrame 列名为扁平化名称以供显示和选择
            df_ex.columns = flat_columns

            st.success("✅ 表头识别成功")
            st.dataframe(df_ex.head())

            st.subheader("字段选择")

            selected_cols = st.multiselect(
                "请选择需要保留的字段：",
                options=flat_columns
            )

            default_ex_name = os.path.splitext(file_extract.name)[0] + "_提取"
            ex_out_name = st.text_input("导出文件名:", value=default_ex_name)

            if selected_cols:
                if st.button("🚀 生成并导出新表"):
                    new_df = df_ex[selected_cols].copy()
                    
                    # 简化逻辑：直接导出扁平化后的单行表头
                    out_ex_io = BytesIO()
                    with pd.ExcelWriter(out_ex_io, engine='openpyxl') as writer:
                        new_df.to_excel(writer, index=False)

                    st.balloons()
                    st.download_button(
                        f"📥 下载 {ex_out_name}.xlsx",
                        out_ex_io.getvalue(),
                        f"{ex_out_name}.xlsx"
                    )
            else:
                st.warning("请至少选择一个字段")

        except Exception as e:
            st.error(f"读取失败: {e}")


# ======================================================
# 底部
# ======================================================
st.divider()
st.caption("💡 所有操作在本地执行。复杂 .xls 文件建议先转为 .xlsx 再使用。")